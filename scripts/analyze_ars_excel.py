#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyze ARS_BSW-ASW_Interface_V1.00.xlsx and produce a concise Markdown report
for preparing V1.01 updates.

Usage:
  venv\Scripts\python scripts\analyze_ars_excel.py \
    --input communication\shared\wia_ars\ARS_BSW-ASW_Interface_V1.00.xlsx \
    --output communication\shared\wia_ars\ARS_V1.01_prep_report.md
"""

import argparse
import sys
from pathlib import Path
from typing import List, Tuple
from openpyxl import load_workbook


API_FUNCTIONS = {
    "ADC": [
        "ShrHWIA_BswAdc_GetPhyValue",
        "ShrHWIA_BswAdc_GetCurrentCalibStatus",
        "ShrHWIA_BswAdc_GetMotorTempErrStatus",
    ],
    "CAN": [
        "ShrHWIA_BswCan_GetMsg",
        "ShrHWIA_BswCan_SetMsg",
        "ShrHWIA_BswCan_GetState_Busoff",
        "ShrHWIA_BswCan_GetState_Timeout",
    ],
    "SYS": [
        "ShrHWIA_BswSys_GetCpuLoad",
        "ShrHWIA_BswSys_SetTargetAxle",
        "ShrHWIA_BswSys_GetResetReason",
    ],
}

STRUCT_NAMES = [
    "typBswCanMsg_MVPC1",
    "typBswCanMsg_RT1_10",
    "typBswCanMsg_RT1_20",
    "typBswCanMsg_RT1_200",
    "typBswCanMsg_ARS1",
    "DebugADC_t",
    "DebugENC_t",
    "DebugIcuStatus_t",
]

PATTERNS = {
    "ADC_CHANNEL": ["BSWADC_CH_"],
    "CAN_INDEX": ["CAN", "RX_", "TX_", "_IDX", "INDEX"],
    "DIO_PIN": ["TP", "DIO"],
}


def infer_cell_dtype(v) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    try:
        from datetime import datetime
        if isinstance(v, datetime):
            return "datetime"
    except Exception:
        pass
    return "string"


def sheet_to_rows(ws, max_cols: int = 64) -> List[List]:
    rows: List[List] = []
    for r in ws.iter_rows(values_only=True):
        row = list(r[:max_cols])
        rows.append(row)
    return rows

def summarize_columns(rows: List[List]) -> List[Tuple[str, str, int]]:
    if not rows:
        return []
    header = rows[0]
    n_cols = len(header)
    cols = [str(c) if c not in (None, "") else f"col_{i+1}" for i, c in enumerate(header)]
    data = rows[1:]
    summary: List[Tuple[str, str, int]] = []
    for ci in range(n_cols):
        col_vals = [row[ci] if ci < len(row) else None for row in data]
        dtype = "string"
        for v in col_vals:
            if v is not None and str(v) != "":
                dtype = infer_cell_dtype(v)
                break
        na = sum(1 for v in col_vals if v is None or (isinstance(v, str) and v.strip() == ""))
        summary.append((cols[ci], dtype, na))
    return summary

def search_token_rows(rows: List[List], token: str) -> bool:
    t = token.lower()
    for row in rows:
        for v in row:
            if v is None:
                continue
            try:
                if t in str(v).lower():
                    return True
            except Exception:
                continue
    return False


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    args = p.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=in_path, data_only=True, read_only=True)
    sheets = wb.sheetnames

    lines: list[str] = []
    lines.append("# ARS V1.01 준비 보고서 (자동 생성)")
    lines.append("")
    lines.append("## 기본 정보 요약")
    lines.append(f"- 총 시트 개수: {len(sheets)}개")
    lines.append("- 시트 목록: " + ", ".join(sheets))
    lines.append("")

    # Detailed per-sheet
    lines.append("## 시트별 상세 분석")
    for s in sheets:
        ws = wb[s]
        rows_data = sheet_to_rows(ws)
        rows = len(rows_data)
        cols = len(rows_data[0]) if rows_data else 0
        lines.append(f"### 시트명: {s}")
        lines.append(f"- 총 행/열 개수: {rows} / {cols}")
        if cols and rows_data:
            lines.append("- 주요 컬럼:")
            for name, dtype, na in summarize_columns(rows_data)[: min(12, cols)]:
                lines.append(f"  * {name}: {dtype}, 결측치 {na}")
        # Sample top-5
        if rows_data:
            header = [str(h) if h not in (None, '') else '' for h in (rows_data[0] if rows_data else [])]
            data_rows = rows_data[1:6]
            if header:
                hdr = header[: min(8, len(header))]
                cols_hdr = " | ".join(hdr)
                sep = " | ".join(["---"] * len(hdr))
                lines.append("- 표본 데이터 (상위 5행):")
                lines.append("| " + cols_hdr + " |")
                lines.append("| " + sep + " |")
                for r in data_rows:
                    row_vals = [str(v)[:64] if v is not None else "" for v in r[: len(hdr)]]
                    lines.append("| " + " | ".join(row_vals) + " |")
        lines.append("")

    # API mapping across all sheets by raw text search
    lines.append("## API 매핑 결과")
    api_found = []
    api_missing = []
    for cat, funcs in API_FUNCTIONS.items():
        for fn in funcs:
            found_one = False
            for s in sheets:
                ws = wb[s]
                rows_data = sheet_to_rows(ws)
                if search_token_rows(rows_data, fn):
                    lines.append(f"- {fn} : {s} 시트 내 기재됨")
                    api_found.append(fn)
                    found_one = True
                    break
            if not found_one:
                api_missing.append(fn)
    if api_missing:
        lines.append("")
        lines.append("### 문서화 누락 API (V1.01 추가 필요)")
        for fn in api_missing:
            lines.append(f"- {fn}")

    # Structure names
    lines.append("")
    lines.append("## 데이터 구조체 문서화 현황")
    for name in STRUCT_NAMES:
        present = False
        for s in sheets:
            ws = wb[s]
            rows_data = sheet_to_rows(ws)
            if search_token_rows(rows_data, name):
                lines.append(f"- {name} : {s} 시트 내 기재됨")
                present = True
                break
        if not present:
            lines.append(f"- {name} : 미확인")

    # Channels/index/pins
    lines.append("")
    lines.append("## 채널 및 인덱스 정의 현황 (키워드 스캔)")
    for label, needles in PATTERNS.items():
        hits_summary = []
        for s in sheets:
            ws = wb[s]
            rows_data = sheet_to_rows(ws)
            if any(search_token_rows(rows_data, tok) for tok in needles):
                hits_summary.append(s)
        if hits_summary:
            lines.append(f"- {label}: 시트 {', '.join(hits_summary)} 에서 발견")
        else:
            lines.append(f"- {label}: 미확인")

    # V1.01 recommendations (template)
    lines.append("")
    lines.append("## V1.01 업데이트 권장사항(초안)")
    lines.append("- P1: SENT 관련 API 문서화 반영, 누락 API 보강")
    lines.append("- P2: 성능 모니터링/시스템 제어 API 단락 확장")
    lines.append("- P3: 테스트/디버그 구조체 정의 보완")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(str(out_path))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
